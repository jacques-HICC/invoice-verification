# app/processing/file_converter.py
import os
import openpyxl
from openpyxl.drawing.image import Image as ExcelImage
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
from reportlab.lib.units import inch
from PIL import Image
import tempfile

class FileConverter:
    """Convert various file types to PDF for processing"""
    
    SUPPORTED_EXTENSIONS = {
        '.xlsx', '.xls',  # Excel
        '.png', '.jpg', '.jpeg', '.tiff', '.bmp',  # Images
        '.txt', '.csv'  # Text files
    }
    
    @staticmethod
    def needs_conversion(filepath: str) -> bool:
        """Check if file needs conversion to PDF"""
        ext = os.path.splitext(filepath)[1].lower()
        return ext in FileConverter.SUPPORTED_EXTENSIONS
    
    @staticmethod
    def convert_to_pdf(input_path: str, output_path: str = None) -> str:
        """
        Convert file to PDF. Returns path to PDF file.
        If output_path is None, creates temp file.
        """
        ext = os.path.splitext(input_path)[1].lower()
        
        if output_path is None:
            temp_dir = os.path.join(os.getcwd(), "temp")
            os.makedirs(temp_dir, exist_ok=True)
            output_path = os.path.join(
                temp_dir, 
                f"{os.path.splitext(os.path.basename(input_path))[0]}.pdf"
            )
        
        if ext in ['.xlsx', '.xls']:
            return FileConverter._excel_to_pdf(input_path, output_path)
        elif ext in ['.png', '.jpg', '.jpeg', '.tiff', '.bmp']:
            return FileConverter._image_to_pdf(input_path, output_path)
        elif ext in ['.txt', '.csv']:
            return FileConverter._text_to_pdf(input_path, output_path)
        else:
            raise ValueError(f"Unsupported file type: {ext}")
    
    @staticmethod
    def _excel_to_pdf(excel_path: str, pdf_path: str) -> str:
        """Convert Excel to PDF using openpyxl + reportlab"""
        wb = openpyxl.load_workbook(excel_path, data_only=True)
        
        c = canvas.Canvas(pdf_path, pagesize=letter)
        width, height = letter
        
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            
            # Sheet title
            c.setFont("Helvetica-Bold", 14)
            c.drawString(50, height - 50, f"Sheet: {sheet_name}")
            
            y_position = height - 100
            c.setFont("Helvetica", 9)
            
            # Extract data from cells
            for row_idx, row in enumerate(sheet.iter_rows(values_only=True), 1):
                if y_position < 50:  # New page if needed
                    c.showPage()
                    y_position = height - 50
                    c.setFont("Helvetica", 9)
                
                row_text = " | ".join([str(cell) if cell is not None else "" for cell in row])
                
                # Truncate long rows
                if len(row_text) > 100:
                    row_text = row_text[:100] + "..."
                
                c.drawString(50, y_position, row_text)
                y_position -= 15
            
            c.showPage()  # New page for next sheet
        
        c.save()
        print(f"✅ Excel converted to PDF: {pdf_path}")
        return pdf_path
    
    @staticmethod
    def _image_to_pdf(image_path: str, pdf_path: str) -> str:
        """Convert image to PDF"""
        c = canvas.Canvas(pdf_path, pagesize=letter)
        width, height = letter
        
        # Open and resize image to fit page
        img = Image.open(image_path)
        img_width, img_height = img.size
        
        # Calculate scaling to fit page (with margins)
        max_width = width - 100
        max_height = height - 100
        
        scale = min(max_width / img_width, max_height / img_height)
        new_width = img_width * scale
        new_height = img_height * scale
        
        # Center image
        x = (width - new_width) / 2
        y = (height - new_height) / 2
        
        c.drawImage(image_path, x, y, width=new_width, height=new_height)
        c.save()
        
        print(f"✅ Image converted to PDF: {pdf_path}")
        return pdf_path
    
    @staticmethod
    def _text_to_pdf(text_path: str, pdf_path: str) -> str:
        """Convert text/CSV file to PDF"""
        with open(text_path, 'r', encoding='utf-8') as f:
            lines = f.readlines()
        
        c = canvas.Canvas(pdf_path, pagesize=letter)
        width, height = letter
        
        c.setFont("Courier", 9)  # Monospace for alignment
        y_position = height - 50
        
        for line in lines:
            if y_position < 50:
                c.showPage()
                y_position = height - 50
                c.setFont("Courier", 9)
            
            # Truncate very long lines
            line = line.rstrip()[:120]
            c.drawString(50, y_position, line)
            y_position -= 12
        
        c.save()
        print(f"✅ Text/CSV converted to PDF: {pdf_path}")
        return pdf_path